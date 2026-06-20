"""POST /api/liquidation/compute — M4. Router de liquidación de prestaciones sociales.

Endpoints:
  POST /compute  → JSON con los valores calculados (determinista)
  POST /export   → archivo .xlsx con la plantilla de liquidación HG (para auditoría y firma)

El formato Excel replica EXACTAMENTE Liquidacion_Andres.xlsx:
  - Hoja "LIQUIDACIÓN" con dos secciones en la misma hoja.
  - Sección 1 (filas 2-61):  liquidación de prestaciones sociales.
  - Sección 2 (filas 64-97): indemnización por terminación.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.security import get_tenant
from app.core.tenancy import TenantContext
from app.api.deps import get_liquidation_service
from app.domain.liquidation.engine import LiquidationInput, LiquidationResult
from app.services.liquidation_service import LiquidationService

router = APIRouter(prefix="/api/liquidation", tags=["liquidation"])

_MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]
_MESES_ES_UPPER = [m.upper() for m in _MESES_ES]
_DIAS_SEMANA = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]


class LiquidationRequest(BaseModel):
    doc_id: str
    monthly_salary: float
    days_worked: int
    vinculo_type: str
    promedio_variable: float = 0.0
    auxilio_transporte: float = 0.0
    dias_pendientes_vacaciones: int = 0
    termination_cause: str = "sin_justa_causa"
    months_remaining_fixed: int = 0
    antiguedad_anios: float = 0.0
    bonificacion: float = 0.0
    # Datos del trabajador y contrato — necesarios para el Excel (los rellena batch_service)
    nombre_trabajador: str = ""
    documento_identidad: str = ""   # "C.C. 94.556.201"
    cargo: str = ""
    centro_costos: str = ""
    nombre_empresa: str = ""
    vinculo_type_label: str = "Término indefinido"
    fecha_inicio: Optional[str] = None      # ISO "2024-03-01"
    fecha_terminacion: Optional[str] = None
    motivo_terminacion: str = "Sin especificar"
    dias_ultimo_periodo: Optional[int] = None


def _build_inp(req: LiquidationRequest) -> LiquidationInput:
    return LiquidationInput(
        salario_basico=req.monthly_salary, days_worked=req.days_worked,
        vinculo_type=req.vinculo_type, termination_cause=req.termination_cause,
        promedio_variable=req.promedio_variable, auxilio_transporte=req.auxilio_transporte,
        dias_pendientes_vacaciones=req.dias_pendientes_vacaciones,
        months_remaining_fixed=req.months_remaining_fixed,
        antiguedad_anios=req.antiguedad_anios, bonificacion=req.bonificacion,
    )


@router.post("/compute")
def compute(
    req: LiquidationRequest,
    ctx: TenantContext = Depends(get_tenant),
    svc: LiquidationService = Depends(get_liquidation_service),
):
    result = svc.compute(ctx, req.doc_id, _build_inp(req))
    return {"doc_id": req.doc_id, "items": result.__dict__, "deterministic": True}


@router.post("/export")
def export_excel(
    req: LiquidationRequest,
    ctx: TenantContext = Depends(get_tenant),
    svc: LiquidationService = Depends(get_liquidation_service),
):
    result = svc.compute(ctx, req.doc_id, _build_inp(req))
    xlsx_bytes = _build_xlsx(req, result)
    nombre = (req.nombre_trabajador or req.doc_id).replace(" ", "_")
    filename = f"liquidacion_{nombre}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Helpers ────────────────────────────────────────────────────────────────────

def _cop(value: float) -> str:
    """1750905 → '$ 1.750.905'  (formato colombiano con puntos)."""
    return "$ {:,}".format(int(round(value))).replace(",", ".")


def _num_to_words(n: int) -> str:
    """Convierte entero a palabras en español (género masculino — pesos)."""
    if n == 0:
        return "cero"
    if n < 0:
        return "menos " + _num_to_words(-n)

    ONES = [
        "", "un", "dos", "tres", "cuatro", "cinco", "seis", "siete",
        "ocho", "nueve", "diez", "once", "doce", "trece", "catorce",
        "quince", "dieciséis", "diecisiete", "dieciocho", "diecinueve", "veinte",
    ]
    VEINTI = [
        "veinte", "veintiún", "veintidós", "veintitrés", "veinticuatro",
        "veinticinco", "veintiséis", "veintisiete", "veintiocho", "veintinueve",
    ]
    TENS = ["", "diez", "veinte", "treinta", "cuarenta", "cincuenta",
            "sesenta", "setenta", "ochenta", "noventa"]
    HUNDS = ["", "cien", "doscientos", "trescientos", "cuatrocientos",
             "quinientos", "seiscientos", "setecientos", "ochocientos", "novecientos"]

    if n >= 1_000_000:
        m, r = divmod(n, 1_000_000)
        base = "un millón" if m == 1 else _num_to_words(m) + " millones"
        return (base + " " + _num_to_words(r)).strip() if r else base

    if n >= 1_000:
        k, r = divmod(n, 1_000)
        base = "mil" if k == 1 else _num_to_words(k) + " mil"
        return (base + " " + _num_to_words(r)).strip() if r else base

    if n >= 100:
        h, r = divmod(n, 100)
        base = "cien" if n == 100 else HUNDS[h]
        return (base + " " + _num_to_words(r)).strip() if r else base

    if n <= 20:
        return ONES[n]

    if n < 30:
        return VEINTI[n - 20]

    t, u = divmod(n, 10)
    return TENS[t] if u == 0 else f"{TENS[t]} y {ONES[u]}"


def _date_iso(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def _month_breakdown(start: date, end: date) -> list[tuple[str, int]]:
    """Días trabajados por mes (convenio 30 días/mes CST), entre start y end inclusive."""
    rows: list[tuple[str, int]] = []
    cur_m, cur_y = start.month, start.year
    while (cur_y, cur_m) <= (end.year, end.month):
        if cur_y == start.year and cur_m == start.month:
            if cur_y == end.year and cur_m == end.month:
                days = end.day - start.day + 1
            else:
                days = 30 - start.day + 1
        elif cur_y == end.year and cur_m == end.month:
            days = end.day
        else:
            days = 30
        rows.append((_MESES_ES_UPPER[cur_m - 1], max(0, days)))
        cur_m += 1
        if cur_m > 12:
            cur_m, cur_y = 1, cur_y + 1
    return rows


def _periodo_str(d_start: date, d_end: date) -> str:
    """Del 01 de enero de 2026 al 04 de mayo de 2026"""
    return (
        f"Del {d_start.day:02d} de {_MESES_ES[d_start.month - 1]} de {d_start.year} "
        f"al {d_end.day:02d} de {_MESES_ES[d_end.month - 1]} de {d_end.year}"
    )


# ── Constructor Excel ───────────────────────────────────────────────────────────

def _build_xlsx(req: LiquidationRequest, result: LiquidationResult) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl no instalado. Ejecute: pip install openpyxl")

    # ── Valores base ──────────────────────────────────────────────────────────
    sal_basico    = req.monthly_salary
    promedio_var  = req.promedio_variable or 0.0
    auxilio       = req.auxilio_transporte or 0.0
    sal_base_liq  = sal_basico + promedio_var + auxilio   # C18
    sal_ordinario = sal_basico + promedio_var              # C76 (sin auxilio)

    fecha_inicio = _date_iso(req.fecha_inicio)
    fecha_term   = _date_iso(req.fecha_terminacion) or date.today()
    year         = fecha_term.year
    dia_term     = fecha_term.day
    mes_term     = _MESES_ES[fecha_term.month - 1]
    mes_term_up  = _MESES_ES_UPPER[fecha_term.month - 1]

    dias_ult = req.dias_ultimo_periodo if req.dias_ultimo_periodo is not None else dia_term

    # Períodos
    # Cesantías: 1-ene-año (o fecha inicio si es el mismo año) → fecha terminación
    if fecha_inicio and fecha_inicio.year == year:
        ces_start = fecha_inicio
    else:
        ces_start = date(year, 1, 1)
    ces_period = _periodo_str(ces_start, fecha_term)
    ces_year_label = f"CESANTÍAS {year}"
    int_ces_year_label = f"INT. CESANTIAS {year}"

    # Prima: semestre corriente
    prima_start = date(year, 1, 1) if fecha_term.month <= 6 else date(year, 7, 1)
    prima_period = _periodo_str(prima_start, fecha_term)

    # Vacaciones: sin registro de disfrute (default)
    vac_period = "Sin registro de disfrute"

    # Adicionales
    sal_periodo = round((sal_basico / 30) * dias_ult)
    aux_periodo = round((auxilio / 30) * dias_ult)
    salud       = round(sal_basico * 0.04)
    pension_ded = round(sal_basico * 0.04)
    total_adic  = sal_periodo + aux_periodo - salud - pension_ded
    neto_pagar  = round(result.total_prestaciones + total_adic)

    # Label fila 36
    sal_adic_label = f"SALARIO DEL 01 AL {dia_term:02d} DE {mes_term_up} DE {year}"

    # Desglose mensual (panel lateral G-H)
    breakdown = _month_breakdown(ces_start, fecha_term)

    # Título hoja
    vinculo_map = {
        "termino_indefinido": "TÉRMINO INDEFINIDO",
        "termino_fijo":        "TÉRMINO FIJO",
        "obra_o_labor":        "OBRA O LABOR",
        "prestacion_servicios":"PRESTACIÓN DE SERVICIOS",
    }
    vinculo_title = vinculo_map.get(req.vinculo_type, "TÉRMINO INDEFINIDO")
    title = f"LIQUIDACIÓN POR TERMINACIÓN DE CONTRATO DE TRABAJO A {vinculo_title}"

    # Texto legal indemnización
    ind_total = int(round(result.indemnizacion))
    ind_letras = _num_to_words(ind_total)
    empresa = req.nombre_empresa or "la empresa"
    term_fecha_str = f"{dia_term:02d} de {mes_term} de {year}"
    # Formato colombiano para el monto: 1750905 → 1.750.905
    ind_monto_fmt = f"{ind_total:,}".replace(",", ".")
    legal_ind = (
        f"Recibo la suma de {ind_letras} pesos m/c "
        f"(${ind_monto_fmt}.oo), a título de indemnización, declarando a {empresa}, "
        "estar a paz y salvo por concepto de indemnizaciones o derechos inciertos "
        "laborales y cualquier tipo de derecho cierto e indiscutible derivado del "
        "contrato de trabajo y la relación laboral habida a la fecha, la cual "
        f"finaliza el día {term_fecha_str}."
    )

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación HG"

    # Anchos de columna (idénticos al template)
    ws.column_dimensions["A"].width = 3.82
    ws.column_dimensions["B"].width = 58.63
    ws.column_dimensions["C"].width = 38.18
    ws.column_dimensions["D"].width = 25.18
    ws.column_dimensions["E"].width = 15.63
    ws.column_dimensions["F"].width = 19.18
    ws.column_dimensions["G"].width = 19.63
    ws.column_dimensions["H"].width = 17.82

    # Altura de filas clave
    for r, h in [
        (2,21),(3,14.25),(5,15),(6,14.25),(7,14.25),(8,14.25),(9,14.25),
        (11,14.25),(12,14.25),(13,14.25),(15,14.25),(16,14.25),(17,14.25),(18,14.25),
        (20,15),(21,15),(23,14.25),(24,15),(26,14.25),(27,15),(29,14.25),(30,14.25),
        (33,14.25),(35,15),(36,14.25),(37,14.25),(38,14.25),(41,14.25),(42,14.25),
        (45,14.25),(48,18),(51,14.25),(52,42),(53,48),(54,14.25),
        (55,18),(56,18),(57,14.25),(58,14.25),(59,14.25),(60,18),(61,15),
        (64,15),(65,15),(66,15),(67,15),(68,15),(69,15),(70,15),(71,15),
        (72,15),(73,15),(74,15),(75,15),(76,15),(77,15),(78,15),(79,15),
        (80,15),(81,15),(82,15),(83,15),(84,15),(85,15),(86,15),(87,15),
        (88,58.5),(89,15),(90,15),(91,15),(92,15),(93,15),(94,15),(95,15),(96,15),(97,15),
    ]:
        ws.row_dimensions[r].height = h

    # ── Estilos base ──────────────────────────────────────────────────────────
    FMT_COP  = '_("$ "* #,##0_);_("$ "* \\(#,##0\\);_("$ "* \\-??_);_(@_)'
    FMT_DATE = 'm/d/yyyy'
    FMT_DATE_LONG = '[$-F800]dddd", "mmmm\\ dd", "yyyy'
    FMT_INT  = '#,##0'

    def _f(bold=False, size=11, color="000000", italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    def _al(h="general", v="bottom", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _set(row, col, value, font=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font = font
        if align: c.alignment = align
        if num_fmt: c.number_format = num_fmt
        return c

    def _merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def _lbl(row, col, text, size=11):
        """Celda de label (negrita)."""
        _set(row, col, text, font=_f(bold=True, size=size))

    def _val(row, col, value, size=11, align=None, fmt=None):
        """Celda de valor (no negrita)."""
        _set(row, col, value, font=_f(size=size), align=align, num_fmt=fmt)

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN 1 — LIQUIDACIÓN (filas 2-61)
    # ════════════════════════════════════════════════════════════════════════

    # Fila 2: Título — B2:E3 merged
    _merge(2, 2, 3, 5)
    _set(2, 2, title,
         font=_f(bold=True, size=14),
         align=_al("center", "center", wrap=True))

    # Filas 6-9: Datos del trabajador
    _lbl(6, 2, "NOMBRE")
    _val(6, 3, req.nombre_trabajador or "")

    _lbl(7, 2, "DOCUMENTO DE IDENTIDAD")
    _val(7, 3, req.documento_identidad or "", fmt=FMT_INT)

    _lbl(8, 2, "CARGO")
    _val(8, 3, req.cargo or "")

    _lbl(9, 2, "CENTRO DE COSTOS")
    _val(9, 3, req.centro_costos or "")

    # Filas 11-13: Fechas y motivo
    _set(11, 2, "FECHA DE INICIO   ", font=_f(bold=True, size=11), align=_al(v="center"))
    _val(11, 3, fecha_inicio, fmt=FMT_DATE)

    _lbl(12, 2, "FECHA TERMINACIÓN CONTRATO")
    _val(12, 3, fecha_term, fmt=FMT_DATE)

    _lbl(13, 2, "MOTIVO DE TERMINACIÓN:")
    _val(13, 3, req.motivo_terminacion)

    # Filas 15-18: Salarios
    _set(15, 2, "SALARIO BASICO", font=_f(size=11))   # NO bold — igual al template
    _val(15, 3, sal_basico, fmt=FMT_COP)
    _val(15, 6, sal_ordinario, fmt='_-"$ "* #,##0_-;"-$ "* #,##0_-;_-"$ "* \\-??_-;_-@_-')

    _set(16, 2, "HORAS EXTRA, RECARGOS Y/O COMISIONES PROMEDIO", font=_f(size=11))
    _val(16, 3, promedio_var, fmt='_-"$ "* #,##0_-;"-$ "* #,##0_-;_-"$ "* \\-??_-;_-@_-')

    _set(17, 2, "AUXILIO DE TRANSPORTE", font=_f(size=11))
    _val(17, 3, auxilio, fmt=FMT_COP)

    _lbl(18, 2, "SALARIO BASE LIQUIDACIÓN")
    _set(18, 3, sal_base_liq,
         font=_f(bold=True, size=11),
         num_fmt=FMT_COP)

    # Panel lateral mes-a-mes (G-H, filas 19-24)
    for i, (mes_name, dias_mes) in enumerate(breakdown[:5]):  # máx 5 meses visibles
        row_g = 19 + i
        _val(row_g, 7, mes_name)
        _val(row_g, 8, dias_mes, fmt="0")
    # Total días (H24)
    _val(24, 8, sum(d for _, d in breakdown), fmt="0")

    # Filas 20-21: CESANTÍAS
    _set(20, 2, ces_year_label, font=_f(bold=True, size=11), align=_al("left", "center"))
    _merge(20, 3, 20, 4)
    _val(20, 3, ces_period, align=_al("center"), fmt=FMT_INT)

    _merge(21, 2, 21, 3)
    _set(21, 2, str(req.days_worked),
         font=_f(size=11), align=_al("right", "center"), num_fmt=FMT_INT)
    _merge(21, 4, 21, 5)
    _set(21, 4, result.cesantias,
         font=_f(bold=True, size=12), align=_al("center"), num_fmt=FMT_COP)

    # Filas 23-24: INT. CESANTIAS
    _lbl(23, 2, int_ces_year_label)
    _merge(23, 3, 23, 4)
    _val(23, 3, ces_period, align=_al("center"), fmt=FMT_INT)

    _merge(24, 2, 24, 3)
    _set(24, 2, str(req.days_worked),
         font=_f(size=11), align=_al("right"), num_fmt=FMT_INT)
    _merge(24, 4, 24, 5)
    _set(24, 4, result.intereses_cesantias,
         font=_f(bold=True, size=12), align=_al("center"), num_fmt=FMT_COP)

    # Filas 26-27: PRIMA DE SERVICIOS
    _lbl(26, 2, "PRIMA DE SERVICIOS")
    _merge(26, 3, 26, 4)
    _val(26, 3, prima_period, align=_al("center"), fmt=FMT_INT)

    _merge(27, 2, 27, 3)
    _set(27, 2, str(req.days_worked),
         font=_f(size=11), align=_al("right", "center"), num_fmt=FMT_INT)
    _merge(27, 4, 27, 5)
    _set(27, 4, result.prima,
         font=_f(bold=True, size=12), align=_al("center"), num_fmt=FMT_COP)

    # Filas 29-30: VACACIONES
    _lbl(29, 2, "VACACIONES")
    _merge(29, 3, 29, 4)
    _val(29, 3, vac_period, align=_al("center"), fmt=FMT_INT)

    _merge(30, 2, 30, 3)
    _set(30, 2, str(req.dias_pendientes_vacaciones),
         font=_f(size=11), align=_al("right"), num_fmt=FMT_INT)
    _merge(30, 4, 30, 5)
    _set(30, 4, result.vacaciones,
         font=_f(bold=True, size=12), align=_al("center", "center"),
         num_fmt='_-"$ "* #,##0_-;"-$ "* #,##0_-;_-"$ "* \\-??_-;_-@_-')

    # Fila 33: TOTAL LIQUIDACIÓN
    _lbl(33, 2, "TOTAL LIQUIDACIÓN")
    _merge(33, 4, 33, 5)
    _set(33, 4, result.total_prestaciones,
         font=_f(bold=True, size=11), align=_al("center"), num_fmt=FMT_COP)

    # Fila 35: ADICIONALES
    _lbl(35, 2, "ADICIONALES")

    # Fila 36: Salario último período
    _set(36, 2, sal_adic_label, font=_f(size=11))
    _set(36, 4, sal_periodo, font=_f(size=11), num_fmt=FMT_COP)

    # Fila 37: Auxilio de transporte pro-rata
    _merge(37, 2, 37, 3)
    _set(37, 2, "AUXILIO DE TRANSPORTE",
         font=_f(size=11), align=_al(wrap=True))
    _set(37, 4, aux_periodo, font=_f(size=11), num_fmt=FMT_COP)

    # Fila 38: Horas extra
    _set(38, 2, "HORAS EXTRA Y RECARGOS",
         font=_f(size=11), align=_al(wrap=True))
    _set(38, 4, 0, font=_f(size=11), num_fmt=FMT_COP)

    # Filas 41-42: Deducciones (rojo)
    _set(41, 2, "(-) SALUD", font=_f(size=11, color="FF0000"))
    _set(41, 4, salud, font=_f(size=11, color="FF0000"), num_fmt=FMT_COP)

    _set(42, 2, "(-) PENSIÓN", font=_f(size=11, color="FF0000"))
    _set(42, 4, pension_ded, font=_f(size=11, color="FF0000"), num_fmt=FMT_COP)

    # Fila 45: Total adicional
    _lbl(45, 2, "TOTAL ADICIONAL")
    _set(45, 4, total_adic, font=_f(bold=True, size=11), num_fmt=FMT_COP)

    # Fila 48: NETO A PAGAR
    _lbl(48, 2, "NETO A PAGAR", size=14)
    _merge(48, 4, 48, 5)
    _set(48, 4, neto_pagar,
         font=_f(bold=True, size=14), align=_al("center"), num_fmt=FMT_COP)

    # Filas 51-53: Texto legal firma (B51:E53 merged)
    _merge(51, 2, 53, 5)
    _set(51, 2, "Recibo valor de liquidación.",
         font=_f(bold=True, size=11), align=_al("left", "top", wrap=True))

    # Fila 54: Encabezados de firma
    _set(54, 2, "Recibí,", font=_f(size=11))
    _set(54, 4, "Aprobó,", font=_f(size=11))

    # Filas 58-59: Nombres firma
    _merge(58, 2, 58, 3)
    _set(58, 2, req.nombre_trabajador or "",
         font=_f(bold=True, size=11))
    _set(59, 2, req.documento_identidad or "",
         font=_f(bold=True, size=11), num_fmt=FMT_INT)
    _set(59, 4, "Representante Legal", font=_f(size=11))

    # Separador / pie sección 1
    _merge(60, 2, 60, 4)
    _merge(61, 2, 61, 5)

    # ════════════════════════════════════════════════════════════════════════
    # SECCIÓN 2 — INDEMNIZACIÓN (filas 64-97)
    # ════════════════════════════════════════════════════════════════════════

    # Fila 64: Título sección
    _merge(64, 2, 64, 4)
    _set(64, 2, "INDEMNIZACIÓN",
         font=_f(bold=True, size=12), align=_al("center", "center"), num_fmt=FMT_INT)

    # Filas 68-76: Datos del trabajador (repetidos para la sección indemnización)
    _lbl(68, 2, "NOMBRE", size=12); _val(68, 3, req.nombre_trabajador or "", size=12)
    _lbl(69, 2, "DOCUMENTO DE IDENTIDAD", size=12)
    _val(69, 3, req.documento_identidad or "", size=12, fmt=FMT_INT)
    _lbl(70, 2, "CARGO", size=12); _val(70, 3, req.cargo or "", size=12, fmt=FMT_INT)
    _lbl(71, 2, "TIPO DE CONTRATO", size=12)
    _val(71, 3, req.vinculo_type_label, size=12, fmt=FMT_INT)

    _lbl(73, 2, "FECHA DE INICIO   ", size=12)
    _val(73, 3, fecha_inicio, size=12, fmt=FMT_DATE_LONG, align=_al(wrap=True))

    _lbl(74, 2, "FECHA TERMINACION CONTRATO", size=12)
    _val(74, 3, fecha_term, size=12, fmt=FMT_DATE_LONG, align=_al(wrap=True))

    _lbl(76, 2, "SALARIO BASICO", size=12)
    _set(76, 3, sal_ordinario,
         font=_f(bold=True, size=12), align=_al("center", "center"), num_fmt=FMT_COP)

    # Filas 80-85: Indemnización
    _merge(80, 2, 84, 2)
    _set(80, 2, "INDEMNIZACIÓN",
         font=_f(bold=True, size=12, color="FF0000"),
         align=_al("center", "center", wrap=True))

    # Fila 81: Primer año — 30 días de salario
    _val(81, 3, "Primer año - 30 días de salario", size=12,
         align=_al(wrap=True))
    _set(81, 4, sal_ordinario,   # 30 días al diario = 1 mes de salario
         font=_f(size=12), align=_al("center", "center"), num_fmt=FMT_COP)

    # Fila 82: Años adicionales (si antigüedad > 1 año)
    anios_add = req.antiguedad_anios - 1 if req.antiguedad_anios > 1 else 0
    if anios_add > 0:
        ind_add = round((sal_ordinario / 30) * 20 * anios_add)
        _val(82, 3,
             f"Año 2 en adelante - 20 días de salario por año ({anios_add:.1f} años)",
             size=12, align=_al(wrap=True))
        _set(82, 4, ind_add,
             font=_f(size=12), align=_al("center", "center"), num_fmt=FMT_COP)

    # Fila 85: Neto indemnización
    _lbl(85, 2, "NETO A PAGAR INDEMNIZACIÓN", size=12)
    _set(85, 4, ind_total,
         font=_f(bold=True, size=12), align=_al("center", "center"), num_fmt=FMT_COP)

    # Fila 88: Texto legal recibo indemnización (B88:D88 merged, justify, wrap)
    _merge(88, 2, 88, 4)
    _set(88, 2, legal_ind,
         font=_f(bold=True, size=11),
         align=_al("justify", "center", wrap=True))

    # Fila 90: Encabezados de firma indemnización
    _set(90, 2, "Recibí,", font=_f(size=12), align=_al("center", "center"))
    _set(90, 4, "Aprobó,", font=_f(size=12), align=_al("center", "center"))

    # Filas 96-97: Nombres firma indemnización
    _set(96, 2, req.nombre_trabajador or "",
         font=_f(bold=True, size=12), align=_al("center", "center"), num_fmt=FMT_COP)
    _set(96, 4, req.nombre_empresa or "",
         font=_f(bold=True, size=12), align=_al("center", "center"), num_fmt=FMT_COP)
    _set(97, 2, req.documento_identidad or "",
         font=_f(bold=True, size=12), align=_al("center"), num_fmt=FMT_INT)

    # ── Hoja 2: Memoria de cálculo (fórmulas + variables + paso a paso) ────────
    _build_memoria_sheet(wb, req, result)

    # ── Guardar ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_memoria_sheet(wb, req: "LiquidationRequest", result: LiquidationResult) -> None:
    """Hoja 'Memoria de Cálculo': muestra cada variable con su fuente, la fórmula
    del CST y el cálculo con los valores reales. Es la transparencia que el jurado
    audita — el mismo número del documento, pero con el porqué a la vista."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Memoria de Cálculo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 2.5

    # Paleta de marca (carbón HG + rojo HG)
    CARBON = "FF2B2B2B"
    HG_RED = "FF801817"
    SOFT   = "FFF3F1EF"
    OK_BG  = "FFEFE9E3"
    WHITE  = "FFFFFFFF"
    GREY   = "FF6B6B6B"
    thin = Side(style="thin", color="FFD9D4CF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cop(v: float) -> str:
        return "$ {:,}".format(int(round(v))).replace(",", ".")

    def cell(r, c, val, *, bold=False, size=10, color="FF2B2B2B", fill=None,
             align="left", wrap=False, border=False, italic=False):
        cc = ws.cell(row=r, column=c, value=val)
        cc.font = Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
        cc.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            cc.fill = PatternFill("solid", fgColor=fill)
        if border:
            cc.border = box
        return cc

    # Valores base (idénticos al motor)
    basico   = req.monthly_salary
    var      = req.promedio_variable or 0.0
    aux      = req.auxilio_transporte or 0.0
    base_liq = basico + var + aux
    base_ord = basico + var
    dias     = req.days_worked
    dpv      = req.dias_pendientes_vacaciones

    # ── Encabezado ────────────────────────────────────────────────────────────
    ws.merge_cells("B2:E2")
    cell(2, 2, "MEMORIA DE CÁLCULO — LIQUIDACIÓN DE PRESTACIONES SOCIALES",
         bold=True, size=13, color=WHITE, fill=CARBON, align="center")
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B3:E3")
    sub = f"{req.nombre_trabajador or '—'}"
    if req.documento_identidad:
        sub += f"  ·  {req.documento_identidad}"
    if req.nombre_empresa:
        sub += f"  ·  {req.nombre_empresa}"
    cell(3, 2, sub, size=10, color=WHITE, fill=HG_RED, align="center")
    ws.row_dimensions[3].height = 18

    # ── Bloque 1: Variables base (con su fuente) ──────────────────────────────
    r = 5
    cell(r, 2, "1 · VARIABLES BASE", bold=True, size=11, color=HG_RED)
    r += 1
    for h, col in (("Variable", 2), ("Valor", 3), ("Fuente", 4)):
        cell(r, col, h, bold=True, size=9, color=WHITE, fill=GREY, align="left", border=True)
    cell(r, 5, "", fill=GREY, border=True)
    r += 1
    variables = [
        ("Salario básico mensual", cop(basico), "Contrato — cláusula de salario"),
        ("Promedio variable (HE/comisiones)", cop(var), "Novedades de nómina (prom. último año)"),
        ("Auxilio de transporte", cop(aux), "Ley — si salario ≤ 2 SMLMV"),
        ("Días trabajados (base 360)", str(dias), "Período liquidado del contrato"),
        ("Días pendientes de vacaciones", str(dpv), "Registro de nómina"),
        ("Base CON auxilio", cop(base_liq), "básico + variable + auxilio"),
        ("Base SIN auxilio", cop(base_ord), "básico + variable (auxilio no es salario)"),
    ]
    for nombre, valor, fuente in variables:
        is_base = nombre.startswith("Base")
        fill = OK_BG if is_base else (SOFT if (r % 2 == 0) else WHITE)
        cell(r, 2, nombre, bold=is_base, size=10, fill=fill, border=True)
        cell(r, 3, valor, bold=is_base, size=10, align="right", fill=fill, border=True)
        cell(r, 4, fuente, size=9, color=GREY, fill=fill, border=True, wrap=True)
        cell(r, 5, "", fill=fill, border=True)
        ws.row_dimensions[r].height = 16
        r += 1

    # ── Bloque 2: Fórmulas y cálculo paso a paso ──────────────────────────────
    r += 1
    cell(r, 2, "2 · FÓRMULAS Y CÁLCULO", bold=True, size=11, color=HG_RED)
    r += 1
    for h, col in (("Concepto", 2), ("Fórmula (norma)", 3), ("Cálculo con valores", 4), ("Resultado", 5)):
        cell(r, col, h, bold=True, size=9, color=WHITE, fill=GREY, border=True)
    r += 1

    # Vacaciones: dos métodos según haya saldo pendiente
    if dpv > 0:
        vac_formula = "base sin aux. ÷ 30 × días pend. (art. 186 CST)"
        vac_calc = f"{cop(base_ord)} ÷ 30 × {dpv}"
    else:
        vac_formula = "base sin aux. × días ÷ 720 (art. 186 CST)"
        vac_calc = f"{cop(base_ord)} × {dias} ÷ 720"

    filas = [
        ("Cesantías", "base × días ÷ 360 (art. 249 CST)",
         f"{cop(base_liq)} × {dias} ÷ 360", result.cesantias),
        ("Intereses sobre cesantías", "cesantías × 12% × días ÷ 360 (art. 99 L.50/90)",
         f"{cop(result.cesantias)} × 12% × {dias} ÷ 360", result.intereses_cesantias),
        ("Prima de servicios", "base × días ÷ 360 (art. 306 CST)",
         f"{cop(base_liq)} × {dias} ÷ 360", result.prima),
        ("Vacaciones", vac_formula, vac_calc, result.vacaciones),
    ]
    for concepto, formula, calc, valor in filas:
        fill = SOFT if (r % 2 == 0) else WHITE
        cell(r, 2, concepto, bold=True, size=10, fill=fill, border=True)
        cell(r, 3, formula, size=9, color=GREY, fill=fill, border=True, wrap=True)
        cell(r, 4, calc, size=9, fill=fill, border=True, wrap=True)
        cell(r, 5, cop(valor), bold=True, size=10, align="right", fill=fill, border=True)
        ws.row_dimensions[r].height = 26
        r += 1

    # Subtotal prestaciones
    cell(r, 2, "TOTAL PRESTACIONES", bold=True, size=11, color=WHITE, fill=CARBON, border=True)
    cell(r, 3, "cesantías + intereses + prima + vacaciones", size=9, color=WHITE, fill=CARBON, border=True, wrap=True)
    cell(r, 4, "", fill=CARBON, border=True)
    cell(r, 5, cop(result.total_prestaciones), bold=True, size=11, color=WHITE, align="right", fill=CARBON, border=True)
    ws.row_dimensions[r].height = 20
    r += 2

    # ── Bloque 3: Indemnización (art. 64) ─────────────────────────────────────
    cell(r, 2, "3 · INDEMNIZACIÓN (art. 64 CST)", bold=True, size=11, color=HG_RED)
    r += 1
    causa = req.termination_cause
    causa_label = {
        "renuncia": "Renuncia voluntaria → sin indemnización",
        "justa_causa": "Despido con justa causa → sin indemnización",
        "mutuo_acuerdo": "Mutuo acuerdo → sin indemnización",
        "transaccion": "Transacción → bonificación acordada",
        "sin_justa_causa": "Despido sin justa causa → indemnización art. 64",
    }.get(causa, causa)
    cell(r, 2, "Motivo de terminación", bold=True, size=10, fill=SOFT, border=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cell(r, 3, causa_label, size=10, fill=SOFT, border=True, wrap=True)
    cell(r, 5, cop(result.indemnizacion), bold=True, size=10, align="right", fill=SOFT, border=True)
    ws.row_dimensions[r].height = 20
    # Fórmula de la indemnización cuando aplica (sin justa causa)
    if causa == "sin_justa_causa" and result.indemnizacion > 0:
        r += 1
        if req.vinculo_type == "termino_fijo":
            ind_formula = (f"salarios que faltaban × meses restantes: "
                           f"{cop(base_ord)} × {req.months_remaining_fixed} meses")
        else:
            anios_add = max(0.0, req.antiguedad_anios - 1)
            ind_formula = (f"30 días + 20 días por año adicional ({anios_add:.1f}): "
                           f"{cop(base_ord)} ÷ 30 × {30 + 20 * anios_add:.0f} días")
        cell(r, 2, "Fórmula (art. 64 CST)", size=9, color=GREY, fill=WHITE, border=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        cell(r, 3, ind_formula, size=9, color=GREY, fill=WHITE, border=True, wrap=True)
        ws.row_dimensions[r].height = 18
    r += 2

    # ── Total final ───────────────────────────────────────────────────────────
    cell(r, 2, "TOTAL A LIQUIDAR", bold=True, size=13, color=WHITE, fill=HG_RED, border=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cell(r, 3, "prestaciones + indemnización", size=10, color=WHITE, fill=HG_RED, border=True)
    cell(r, 5, cop(result.total), bold=True, size=13, color=WHITE, align="right", fill=HG_RED, border=True)
    ws.row_dimensions[r].height = 28
    r += 2

    # Nota de confiabilidad
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    cell(r, 2, "Cálculo determinista: las mismas variables producen siempre el mismo "
               "resultado (función pura, sin IA). Constantes laborales versionadas por año "
               "(SMLMV, interés cesantías 12%, base 360 días).",
         size=8, color=GREY, italic=True, wrap=True)
    ws.row_dimensions[r].height = 30
